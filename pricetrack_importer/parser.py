"""
Parser de exports PriceTrack em formato `.md` (markdown table) ou `.xlsx`.

Colunas (na ordem do export):
    collectionDate | brand | sku | title | marketplace | seller |
    MIN PRICE | AVG PRICE | MODE PRICE | MAX PRICE

Quirks tratados aqui:

1. O campo `title` pode conter o caractere `|`, gerando linhas com NF>11
   ao splittar ingenuamente. Resolvemos ancorando da direita pra esquerda:
   as 6 Ăşltimas colunas (marketplace, seller, MIN, AVG, MODE, MAX) e as 3
   primeiras (collectionDate, brand, sku) sĂŁo confiĂĄveis; o que sobra no
   meio ĂŠ o tĂ­tulo, rejunte com `|`.

2. Linhas de metadados (cabeĂ§alho da tabela, separador `|---|`,
   `Filtros aplicados`, `Total`) sĂŁo identificadas e filtradas mais
   adiante no `validator`.

3. `.xlsx` ĂŠ lido com `openpyxl` em modo read_only â nĂŁo carrega tudo em
   RAM (importante no OptiPlex 3020M de 6GB).
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Generator, List, Optional


# Ordem oficial das colunas no export
EXPECTED_COLUMNS = [
    "collectionDate",
    "brand",
    "sku",
    "title",
    "marketplace",
    "seller",
    "MIN PRICE",
    "AVG PRICE",
    "MODE PRICE",
    "MAX PRICE",
]

# Aliases aceitos (PriceTrack pode mudar capitalizaĂ§ĂŁo)
_COLUMN_ALIASES = {
    "collection date": "collectionDate",
    "collectiondate": "collectionDate",
    "min price": "MIN PRICE",
    "avg price": "AVG PRICE",
    "mode price": "MODE PRICE",
    "max price": "MAX PRICE",
    "min": "MIN PRICE",
    "avg": "AVG PRICE",
    "mode": "MODE PRICE",
    "max": "MAX PRICE",
}


class ParseError(Exception):
    """Erro irrecuperĂĄvel de parsing (arquivo malformado)."""


def parse_file(path: str | Path) -> Generator[Dict[str, str], None, None]:
    """
    Dispatcher por extensĂŁo. Yielda dicts com as 10 colunas + `_line_no`.

    Linhas invĂĄlidas (metadata, separador) NĂO sĂŁo filtradas aqui â passam
    adiante pro `validator`. Esta camada sĂł lida com a forma do arquivo.
    """
    p = Path(path)
    if not p.exists():
        raise ParseError(f"Arquivo nĂŁo encontrado: {p}")

    ext = p.suffix.lower()
    if ext == ".md":
        yield from _parse_markdown(p)
    elif ext in {".xlsx", ".xlsm"}:
        yield from _parse_xlsx(p)
    else:
        raise ParseError(f"ExtensĂŁo nĂŁo suportada: {ext}")


# ----------------------- Markdown --------------------------------------- #


def _parse_markdown(path: Path) -> Generator[Dict[str, str], None, None]:
    """LĂŞ a tabela markdown linha a linha (streaming, sem carregar tudo em RAM)."""
    with open(path, "r", encoding="utf-8") as fh:
        for line_no, raw in enumerate(fh, start=1):
            line = raw.rstrip("\n").rstrip("\r")
            if not line.strip():
                continue
            cells = _split_pipe_row(line)
            if cells is None:
                # NĂŁo ĂŠ uma linha de tabela (texto livre, metadata) â devolve
                # uma linha "vazia" pra contagem; o validator vai rejeitar.
                yield {"_line_no": str(line_no), "collectionDate": "", "_raw": line}
                continue

            row = _row_from_cells(cells, line_no)
            if row is not None:
                yield row
            else:
                # Linha de tabela com formato inesperado â devolve com flag
                yield {
                    "_line_no": str(line_no),
                    "collectionDate": "",
                    "_raw": line,
                    "_unparseable": "1",
                }


def _split_pipe_row(line: str) -> Optional[List[str]]:
    """
    Divide uma linha de tabela markdown por `|`, removendo elementos vazios
    de borda. Devolve None se a linha nĂŁo parece pertencer Ă  tabela.
    """
    if "|" not in line:
        return None

    parts = line.split("|")

    # Tabela markdown costuma ter `|` nas bordas â primeiro e Ăşltimo vĂŞm vazios
    if parts and parts[0].strip() == "":
        parts = parts[1:]
    if parts and parts[-1].strip() == "":
        parts = parts[:-1]

    # Linha separadora `|---|---|...` â todas as cĂŠlulas sĂŁo `---` ou `:---:`
    if all(c.strip().replace(":", "").replace("-", "") == "" and c.strip() for c in parts):
        return None

    # Strip de cada cĂŠlula
    return [c.strip() for c in parts]


def _row_from_cells(cells: List[str], line_no: int) -> Optional[Dict[str, str]]:
    """
    Mapeia cĂŠlulas para o dict final, lidando com pipe-no-tĂ­tulo via
    ancoragem rightmost. Espera no mĂ­nimo 10 colunas (11+ se houver pipe
    extra no tĂ­tulo).
    """
    n = len(cells)

    # CabeĂ§alho: linha onde a primeira cĂŠlula ĂŠ "collectionDate" (caso-insensitivo)
    if n >= 10 and cells[0].lower() in {"collectiondate", "collection date"}:
        return {
            "_line_no": str(line_no),
            "collectionDate": "",
            "_raw": " | ".join(cells),
            "_is_header": "1",
        }

    if n < 10:
        return None  # linha quebrada, deixa o validator rejeitar

    # Ancoragem rightmost: Ăşltimas 4 colunas sĂŁo MIN/AVG/MODE/MAX
    # antes delas: seller, marketplace
    # antes delas: tudo entre [3:-6] ĂŠ o title (junta com "|")
    # antes delas (0..3): collectionDate, brand, sku
    head = cells[:3]
    tail = cells[-6:]  # marketplace, seller, MIN, AVG, MODE, MAX
    title_parts = cells[3:-6]
    title = " | ".join(p for p in title_parts if p) if title_parts else ""

    return {
        "_line_no": str(line_no),
        "collectionDate": head[0],
        "brand": head[1],
        "sku": head[2],
        "title": title,
        "marketplace": tail[0],
        "seller": tail[1],
        "MIN PRICE": tail[2],
        "AVG PRICE": tail[3],
        "MODE PRICE": tail[4],
        "MAX PRICE": tail[5],
    }


# ----------------------- XLSX ------------------------------------------- #


def _parse_xlsx(path: Path) -> Generator[Dict[str, str], None, None]:
    """
    LĂŞ .xlsx em modo read_only via openpyxl (streaming).

    Detecta a linha de cabeĂ§alho buscando uma row que contenha
    `collectionDate` (caso-insensitivo).
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ParseError(
            "openpyxl nĂŁo estĂĄ instalado. Rode `pip install openpyxl` "
            "para suportar exports .xlsx."
        ) from e

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active

    header_idx: Optional[Dict[str, int]] = None
    line_no = 0
    date_col_pos: Optional[int] = None
    for row in ws.iter_rows(values_only=True):
        line_no += 1
        if row is None:
            continue
        cells = [_xlsx_cell_to_str(v) for v in row]

        if header_idx is None:
            mapping = _try_build_header(cells)
            if mapping is not None:
                header_idx = mapping
                date_col_pos = mapping.get("collectionDate")
            continue

        if all(c == "" for c in cells):
            continue

        # Excel grava `collectionDate` como data nativa â openpyxl devolve
        # datetime/date. `_xlsx_cell_to_str` jĂĄ cobre todas as colunas, mas
        # garantimos aqui o formato M/D/YY que o validator espera, mesmo
        # quando a cĂŠlula vier como nĂşmero serial ou outro tipo exĂłtico.
        if date_col_pos is not None and 0 <= date_col_pos < len(row):
            cells[date_col_pos] = _coerce_pricetrack_date(row[date_col_pos])

        out: Dict[str, str] = {"_line_no": str(line_no)}
        for col_name, col_pos in header_idx.items():
            out[col_name] = cells[col_pos] if col_pos < len(cells) else ""
        yield out

    wb.close()


def _xlsx_cell_to_str(v: Any) -> str:
    """
    Converte cĂŠlula do openpyxl para string preservando datas no formato
    M/D/YY que o validator do PriceTrack espera.

    openpyxl com `data_only=True` devolve `datetime`/`date` quando a cĂŠlula
    foi gravada como data nativa no Excel. `str()` ingĂŞnuo produz
    `"2026-05-27 00:00:00"`, que NĂO casa com o regex M/D/YY â todas as
    linhas viram METADATA e nada entra no Supabase.
    """
    if v is None:
        return ""
    if isinstance(v, datetime):
        return f"{v.month}/{v.day}/{v.year % 100:02d}"
    if isinstance(v, date):
        return f"{v.month}/{v.day}/{v.year % 100:02d}"
    return str(v).strip()


def _coerce_pricetrack_date(v: Any) -> str:
    """Garante formato M/D/YY para a coluna collectionDate."""
    if isinstance(v, (datetime, date)):
        return _xlsx_cell_to_str(v)
    return str(v).strip() if v is not None else ""


def _try_build_header(cells: List[str]) -> Optional[Dict[str, int]]:
    """
    Tenta construir o mapeamento {nome_coluna: Ă­ndice} a partir de uma row.

    Devolve None se a row nĂŁo parece ser cabeĂ§alho.
    """
    normalized = [_normalize_header_cell(c) for c in cells]
    found: Dict[str, int] = {}
    for i, h in enumerate(normalized):
        if h in EXPECTED_COLUMNS:
            found[h] = i
    # CritĂŠrio: pelo menos as 4 colunas de preĂ§o + collectionDate apareceram
    required = {"collectionDate", "MIN PRICE", "AVG PRICE", "MODE PRICE", "MAX PRICE"}
    if required.issubset(found):
        # Preenche colunas faltantes (caso bizarro) com Ă­ndice -1
        for col in EXPECTED_COLUMNS:
            found.setdefault(col, -1)
        return found
    return None


def _normalize_header_cell(raw: str) -> str:
    """Converte uma cĂŠlula de cabeĂ§alho para o nome canĂ´nico de coluna."""
    s = raw.strip()
    if not s:
        return s
    low = s.lower()
    if low in _COLUMN_ALIASES:
        return _COLUMN_ALIASES[low]
    # Match case-insensitive contra EXPECTED_COLUMNS
    for col in EXPECTED_COLUMNS:
        if col.lower() == low:
            return col
    return s
